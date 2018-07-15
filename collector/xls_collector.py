from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from collector.models import Character,WeaponRef,SkillRef,ArmorRef,BeneficeAfflictionRef
from datetime import datetime 
from collector.fs_fics7 import minmax_from_dc
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import load_workbook

def colrow(c,r):
  return '%s%d'%(get_column_letter(c),r)

def export_header(ws,data):
  """ export the header of a set """
  num,r = 1,1
  cell = '%s%d'%(get_column_letter(num),r)
  ws[cell].font = Font(name='Roboto',color='8040C0', bold=True)
  r += 1
  for num,d in enumerate(data,start=1):
    ws.cell(column=num, row=r, value=data[d]['title'])
    ws.column_dimensions[get_column_letter(num)].width = data[d]['width']
    cell = '%s%d'%(get_column_letter(num),r)
    ws[cell].font = Font(name='Roboto',color='8040C0', bold=True, size=9)
    ws[cell].fill = PatternFill(fill_type='solid', fgColor='C0C0C0')

def export_row(ws, data, ch, r):
  """ Export a row from a set """
  for num,d in enumerate(data,start=1):
    ws.cell(column=num, row=r, value='%s'%(getattr(ch,data[d]['attribute'])))
    
def export_to_xls(filename='dramatis_personae.xlsx'):
  """ XLS extraction of the Characters """
  wb = Workbook()
  dest_filename = filename
  ws = wb.active
  # Abstract
  ws.title = 'Abstract'
  ws.column_dimensions['A'].width = 40
  ws.column_dimensions['B'].width = 30
  ws.cell(column=1, row=2, value='Source')
  ws.cell(column=2, row=2, value='Dramatis Personae Collector')
  ws.cell(column=1, row=3, value='Version')
  ws.cell(column=2, row=3, value='0.5')
  ws.cell(column=1, row=4, value='Release date')
  ws.cell(column=2, row=4, value='%s'%(datetime.now()))
  # Characters
  ws = wb.create_sheet('Characters')
  h = {
    '1':{'title':'Name','attribute':'full_name','width':40},
    '2':{'title':'RID','attribute':'rid','width':30},
    '3':{'title':'Entrance','attribute':'entrance','width':40},
    '4':{'title':'Alliance','attribute':'alliance','width':40},
    '5':{'title':'Rank','attribute':'rank','width':30},
    '6':{'title':'Gender','attribute':'gender','width':10},
    '7':{'title':'Species/Race','attribute':'species','width':20},
    '8':{'title':'Caste','attribute':'caste','width':30},
    '9':{'title':'Birthdate','attribute':'birthdate','width':10},
    '10':{'title':'Height','attribute':'height','width':10},
    '11':{'title':'Weight','attribute':'weight','width':10},
    '12':{'title':'STR','attribute':'PA_STR','width':10},
    '13':{'title':'CON','attribute':'PA_CON','width':10},
    '14':{'title':'BOD','attribute':'PA_BOD','width':10},
    '15':{'title':'MOV','attribute':'PA_MOV','width':10},
    '16':{'title':'INT','attribute':'PA_INT','width':10},
    '17':{'title':'WIL','attribute':'PA_WIL','width':10},
    '18':{'title':'TEM','attribute':'PA_TEM','width':10},
    '19':{'title':'PRE','attribute':'PA_PRE','width':10},
    '20':{'title':'TEC','attribute':'PA_TEC','width':10},
    '21':{'title':'REF','attribute':'PA_REF','width':10},
    '22':{'title':'AGI','attribute':'PA_AGI','width':10},
    '23':{'title':'AWA', 'attribute':'PA_AWA','width':10},
  }  
  ws.cell(column=1, row=1, value='Dramatis Personae')
  character_items = Character.objects.all().order_by('full_name')
  export_header(ws,h)
  cnt = 3
  for c in character_items:
    export_row(ws,h,c,cnt)
    cnt += 1

  # Weapons
  ws = wb.create_sheet('Weapons_References')
  h = {
    '1':{'title':'Ref','attribute':'reference','width':40},
    '2':{'title':'Cat','attribute':'category','width':10},
    '3':{'title':'WA','attribute':'weapon_accuracy','width':5},
    '4':{'title':'CO','attribute':'conceilable','width':5},
    '5':{'title':'AV','attribute':'availability','width':5},
    '6':{'title':'DC','attribute':'damage_class','width':15},
    '7':{'title':'cal.','attribute':'caliber','width':15},    
    '8':{'title':'STR','attribute':'str_min','width':5},
    '9':{'title':'RoF','attribute':'rof','width':5},
    '10':{'title':'Clip','attribute':'clip','width':5},
    '11':{'title':'RNG','attribute':'rng','width':5},
    '12':{'title':'RE','attribute':'rel','width':5},
    '13':{'title':'Cost','attribute':'cost','width':10},
    '14':{'title':'Description','attribute':'description','width':40},
  }
  ws.cell(column=1, row=1, value='Dramatis Personae')
  weaponref_items = WeaponRef.objects.all().order_by('category','damage_class')
  export_header(ws,h)
  cnt = 3
  for c in weaponref_items:
    export_row(ws,h,c,cnt)
    cnt += 1

  # SkillRef
  ws = wb.create_sheet('Skills_References')
  h = {
    '1':{'title':'Ref','attribute':'reference','width':30},
    '2':{'title':'Root','attribute':'is_root','width':10},
    '3':{'title':'Speciality','attribute':'is_speciality','width':10},
    '4':{'title':'Category','attribute':'category','width':10},
    '5':{'title':'Linked To','attribute':'linked_to','width':20},
  }
  ws.cell(column=1, row=1, value='Dramatis Personae')
  skillref_items = SkillRef.objects.all().order_by('reference','is_root')
  export_header(ws,h)
  cnt = 3
  for c in skillref_items:
    export_row(ws,h,c,cnt)
    cnt += 1

  # ArmorRef
  ws = wb.create_sheet('Armors_References')
  h = {
    '1':{'title':'Ref','attribute':'reference','width':30},
    '2':{'title':'Category','attribute':'category','width':10},    
    '3':{'title':'Head','attribute':'head','width':5},
    '4':{'title':'Torso','attribute':'torso','width':5},
    '5':{'title':'LeftArm','attribute':'left_arm','width':5},
    '6':{'title':'RightArm','attribute':'right_arm','width':5},
    '7':{'title':'LeftLeg','attribute':'left_leg','width':5},
    '8':{'title':'RightLeg','attribute':'right_leg','width':5},
    '9':{'title':'SP','attribute':'stopping_power','width':5},
    '10':{'title':'Cost','attribute':'cost','width':10},
    '11':{'title':'EV','attribute':'encumbrance','width':5},
    '12':{'title':'Description','attribute':'description','width':30},
  }
  ws.cell(column=1, row=1, value='Dramatis Personae')
  armorref_items = ArmorRef.objects.all().order_by('reference','category')
  export_header(ws,h)
  cnt = 3
  for c in armorref_items:
    export_row(ws,h,c,cnt)
    cnt += 1


  # BeneficeAfflictionRef
  ws = wb.create_sheet('Benefices_Afflicitions_References')
  h = {
    '1':{'title':'Ref','attribute':'reference','width':30},
    '2':{'title':'Value','attribute':'value','width':5},    
    '3':{'title':'Category','attribute':'category','width':10},
    '4':{'title':'Description','attribute':'description','width':30},
    '5':{'title':'ID','attribute':'id','width':5},
  }
  ws.cell(column=1, row=1, value='Dramatis Personae')
  beneficeafflictionref_items = BeneficeAfflictionRef.objects.all().order_by('reference','-value','category')
  export_header(ws,h)
  cnt = 3
  for c in beneficeafflictionref_items:
    export_row(ws,h,c,cnt)
    cnt += 1

  # And save everything
  wb.save(filename = dest_filename)

def update_from_xls(filename='dramatis_personae.xlsx'):
  """
    This is not a real 'import', as we only update some refs from the database.
    No isoprod behavior db <-> xls has to be expected here. THIS IS NO RESTORE !!!
  """
  wb = load_workbook(filename)
  ws = wb['Benefices_Afflicitions_References']
  if ws != None:
    cnt = 3
    #bar = BeneficeAfflictionRef.objects.all().delete()
    while True:
      if ws[colrow(1,cnt)].value is None:
        break
      else:
        print(ws[colrow(1,cnt)].value)
        print(ws[colrow(2,cnt)].value)
        print(ws[colrow(3,cnt)].value)
        print(ws[colrow(4,cnt)].value)
        print(ws[colrow(5,cnt)].value)
        if int(ws[colrow(5,cnt)].value) == 0:
          bar = None
        else:
          bar = BeneficeAfflictionRef.objects.get(id=int(ws[colrow(5,cnt)].value))
        if bar is None:
          bar = BeneficeAfflictionRef(reference=ws[colrow(1,cnt)].value)
        bar.reference = ws[colrow(1,cnt)].value
        bar.value = int(ws[colrow(2,cnt)].value)
        bar.category = ws[colrow(3,cnt)].value
        if ws[colrow(4,cnt)].value is None:
          desc = ""
        else:
          desc = ws[colrow(4,cnt)].value 
        bar.description = desc
        bar.save()
        print(bar)
        print("")
        cnt += 1
