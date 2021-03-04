'''
 ╔╦╗╔═╗  ╔═╗┌─┐┬  ┬  ┌─┐┌─┐┌┬┐┌─┐┬─┐
  ║║╠═╝  ║  │ ││  │  ├┤ │   │ │ │├┬┘
 ═╩╝╩    ╚═╝└─┘┴─┘┴─┘└─┘└─┘ ┴ └─┘┴└─
'''
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from collector.models.character import Character
from collector.models.weapon import WeaponRef
from collector.models.skill import SkillRef
from collector.models.armor import ArmorRef
from collector.models.benefice_affliction import BeneficeAfflictionRef
from datetime import datetime
from collector.utils.fs_fics7 import minmax_from_dc
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import load_workbook
from collector.utils.fics_references import RELEASE


def colrow(c, r):
    return '%s%d' % (get_column_letter(c), r)


def export_header(ws, data):
    """ export the header of a set """
    num, r = 1, 1
    cell = '%s%d' % (get_column_letter(num), r)
    ws[cell].font = Font(name='Eczar ExtraBold', color='8040C0', bold=True)
    r += 1
    for num, d in enumerate(data, start=1):
        ws.cell(column=num, row=r, value=data[d]['title'])
        ws.column_dimensions[get_column_letter(num)].width = data[d]['width']
        cell = '%s%d' % (get_column_letter(num), r)
        ws[cell].font = Font(name='Work Sans Regular', color='8040C0', bold=True, size=9)
        ws[cell].fill = PatternFill(fill_type='solid', fgColor='C0C0C0')


def export_row(ws, data, ch, r):
    """ Export a row from a set """
    for num, dx in enumerate(data, start=1):
        the_field = data[dx]['attribute']
        field_type = ch._meta.get_field(the_field).get_internal_type()
        val = getattr(ch, the_field)
        if field_type == 'ForeignKey':
            related_model = str(self._meta.get_field(the_field).related_model)
            if related_model == "<class 'cartograph.models.fics_models.Specie'>":
                data = Specie.objects.filter(pk=val).first().specie
            # elif related_model == "<class 'cartograph.models.fics_models.Role'>":
            #   data = Role.objects.filter(pk=val).first().reference
            # elif related_model == "<class 'cartograph.models.fics_models.Profile'>":
            #   data = Profile.objects.filter(pk=val).first().reference
            else:
                data = 'Unknown'
        else:
            data = val
        ws.cell(column=num, row=r, value='%s' % (data))


def export_to_xls(filename='dramatis_personae.xlsx'):
    """ XLS extraction of the Characters """
    wb = Workbook()
    dest_filename = filename
    ws = wb.active
    # Abstract
    ws.title = 'Abstract'
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 30
    ws.cell(column=1, row=2, value='Source')
    ws.cell(column=2, row=2, value='Dramatis Personae Collector')
    ws.cell(column=1, row=3, value='Version')
    ws.cell(column=2, row=3, value=str(RELEASE))
    ws.cell(column=1, row=4, value='Release date')
    ws.cell(column=2, row=4, value='%s' % (datetime.now()))
    # Characters
    ws = wb.create_sheet('Characters')
    h = {
        '1': {'title': 'Name', 'attribute': 'full_name', 'width': 40},
        '2': {'title': 'RID', 'attribute': 'rid', 'width': 30},
        '3': {'title': 'Entrance', 'attribute': 'entrance', 'width': 40},
        '4': {'title': 'Alliance', 'attribute': 'alliance', 'width': 40},
        '5': {'title': 'Rank', 'attribute': 'rank', 'width': 30},
        '6': {'title': 'Gender', 'attribute': 'gender', 'width': 10},
        '7': {'title': 'Specie/Race', 'attribute': 'specie', 'width': 20},
        '8': {'title': 'Caste', 'attribute': 'caste', 'width': 30},
        '9': {'title': 'Birthdate', 'attribute': 'birthdate', 'width': 10},
        '10': {'title': 'Age', 'attribute': 'age', 'width': 10},
        '11': {'title': 'Height', 'attribute': 'height', 'width': 10},
        '12': {'title': 'Weight', 'attribute': 'weight', 'width': 10},
        '13': {'title': 'STR', 'attribute': 'PA_STR', 'width': 10},
        '14': {'title': 'CON', 'attribute': 'PA_CON', 'width': 10},
        '15': {'title': 'BOD', 'attribute': 'PA_BOD', 'width': 10},
        '16': {'title': 'MOV', 'attribute': 'PA_MOV', 'width': 10},
        '17': {'title': 'INT', 'attribute': 'PA_INT', 'width': 10},
        '18': {'title': 'WIL', 'attribute': 'PA_WIL', 'width': 10},
        '19': {'title': 'TEM', 'attribute': 'PA_TEM', 'width': 10},
        '20': {'title': 'PRE', 'attribute': 'PA_PRE', 'width': 10},
        '21': {'title': 'TEC', 'attribute': 'PA_TEC', 'width': 10},
        '22': {'title': 'REF', 'attribute': 'PA_REF', 'width': 10},
        '23': {'title': 'AGI', 'attribute': 'PA_AGI', 'width': 10},
        '24': {'title': 'AWA', 'attribute': 'PA_AWA', 'width': 10},
        '25': {'title': 'Recovery', 'attribute': 'SA_REC', 'width': 10},
        '26': {'title': 'Stamina', 'attribute': 'SA_STA', 'width': 10},
        '27': {'title': 'Endurance', 'attribute': 'SA_END', 'width': 10},
        '28': {'title': 'Stun Save', 'attribute': 'SA_STU', 'width': 10},
        '29': {'title': 'Resistance', 'attribute': 'SA_RES', 'width': 10},
        '30': {'title': 'Damage', 'attribute': 'SA_DMG', 'width': 10},
        '31': {'title': 'Tolerance', 'attribute': 'SA_TOL', 'width': 10},
        '32': {'title': 'Humanity', 'attribute': 'SA_HUM', 'width': 10},
        '33': {'title': 'Passion', 'attribute': 'SA_PAS', 'width': 10},
        '34': {'title': 'Wyrd', 'attribute': 'SA_WYR', 'width': 10},
        '35': {'title': 'Speed', 'attribute': 'SA_SPD', 'width': 10},
        '36': {'title': 'Run', 'attribute': 'SA_RUN', 'width': 10},
        '37': {'title': 'Visible', 'attribute': 'is_visible', 'width': 5},
        '38': {'title': 'Exportable', 'attribute': 'is_exportable', 'width': 5},
        '39': {'title': 'Dead', 'attribute': 'is_dead', 'width': 5},
        '40': {'title': 'Locked', 'attribute': 'is_locked', 'width': 5},
    }
    ws.cell(column=1, row=1, value='Dramatis Personae')
    character_items = Character.objects.all().order_by('full_name')
    export_header(ws, h)
    cnt = 3
    for c in character_items:
        export_row(ws, h, c, cnt)
        cnt += 1

    # Weapons
    ws = wb.create_sheet('Weapons_References')
    h = {
        '1': {'title': 'Ref', 'attribute': 'reference', 'width': 40},
        '2': {'title': 'Cat', 'attribute': 'category', 'width': 10},
        '3': {'title': 'WA', 'attribute': 'weapon_accuracy', 'width': 5},
        '4': {'title': 'CO', 'attribute': 'conceilable', 'width': 5},
        '5': {'title': 'AV', 'attribute': 'availability', 'width': 5},
        '6': {'title': 'DC', 'attribute': 'damage_class', 'width': 15},
        '7': {'title': 'cal.', 'attribute': 'caliber', 'width': 15},
        '8': {'title': 'STR', 'attribute': 'str_min', 'width': 5},
        '9': {'title': 'RoF', 'attribute': 'rof', 'width': 5},
        '10': {'title': 'Clip', 'attribute': 'clip', 'width': 5},
        '11': {'title': 'RNG', 'attribute': 'rng', 'width': 5},
        '12': {'title': 'RE', 'attribute': 'rel', 'width': 5},
        '13': {'title': 'Cost', 'attribute': 'cost', 'width': 10},
        '14': {'title': 'Description', 'attribute': 'description', 'width': 40},
    }
    ws.cell(column=1, row=1, value='Dramatis Personae')
    weaponref_items = WeaponRef.objects.all().order_by('category', 'damage_class')
    export_header(ws, h)
    cnt = 3
    for c in weaponref_items:
        export_row(ws, h, c, cnt)
        cnt += 1

    # SkillRef
    ws = wb.create_sheet('Skills_References')
    h = {
        '1': {'title': 'Ref', 'attribute': 'reference', 'width': 30},
        '2': {'title': 'Root', 'attribute': 'is_root', 'width': 10},
        '3': {'title': 'Speciality', 'attribute': 'is_speciality', 'width': 10},
        '4': {'title': 'Linked To', 'attribute': 'linked_to', 'width': 20},
        '5': {'title': 'Group', 'attribute': 'group', 'width': 10},
    }
    ws.cell(column=1, row=1, value='Dramatis Personae')
    skillref_items = SkillRef.objects.all().order_by('is_speciality', 'reference')
    export_header(ws, h)
    cnt = 3
    for c in skillref_items:
        export_row(ws, h, c, cnt)
        cnt += 1

    # ArmorRef
    ws = wb.create_sheet('Armors_References')
    h = {
        '1': {'title': 'Ref', 'attribute': 'reference', 'width': 30},
        '2': {'title': 'Category', 'attribute': 'category', 'width': 10},
        '3': {'title': 'Head', 'attribute': 'head', 'width': 5},
        '4': {'title': 'Torso', 'attribute': 'torso', 'width': 5},
        '5': {'title': 'LeftArm', 'attribute': 'left_arm', 'width': 5},
        '6': {'title': 'RightArm', 'attribute': 'right_arm', 'width': 5},
        '7': {'title': 'LeftLeg', 'attribute': 'left_leg', 'width': 5},
        '8': {'title': 'RightLeg', 'attribute': 'right_leg', 'width': 5},
        '9': {'title': 'SP', 'attribute': 'stopping_power', 'width': 5},
        '10': {'title': 'Cost', 'attribute': 'cost', 'width': 10},
        '11': {'title': 'EV', 'attribute': 'encumbrance', 'width': 5},
        '12': {'title': 'Description', 'attribute': 'description', 'width': 30},
    }
    ws.cell(column=1, row=1, value='Dramatis Personae')
    armorref_items = ArmorRef.objects.all().order_by('reference', 'category')
    export_header(ws, h)
    cnt = 3
    for c in armorref_items:
        export_row(ws, h, c, cnt)
        cnt += 1

    # BeneficeAfflictionRef
    ws = wb.create_sheet('Benefices_Afflicitions_References')
    h = {
        '1': {'title': 'Ref', 'attribute': 'reference', 'width': 30},
        '2': {'title': 'Value', 'attribute': 'value', 'width': 5},
        '3': {'title': 'Category', 'attribute': 'category', 'width': 10},
        '4': {'title': 'Description', 'attribute': 'description', 'width': 60},
        '5': {'title': 'ID', 'attribute': 'id', 'width': 5},
        '6': {'title': 'Source', 'attribute': 'source', 'width': 20},
    }
    ws.cell(column=1, row=1, value='Dramatis Personae')
    beneficeafflictionref_items = BeneficeAfflictionRef.objects.all().order_by('reference', '-value', 'category')
    export_header(ws, h)
    cnt = 3
    for c in beneficeafflictionref_items:
        export_row(ws, h, c, cnt)
        cnt += 1

    # Characters ==> For PC
    ws = wb.create_sheet('Players Characters Catalogue')
    h = {
        '1': {'title': 'Name', 'attribute': 'full_name', 'width': 40},
        '2': {'title': 'Alliance', 'attribute': 'alliance', 'width': 40},
        '3': {'title': 'Dead', 'attribute': 'is_dead', 'width': 10},
        '4': {'title': 'Player', 'attribute': 'player', 'width': 20},
        '5': {'title': 'Rank', 'attribute': 'rank', 'width': 20},
        '6': {'title': 'Gender', 'attribute': 'gender', 'width': 10},
        '7': {'title': 'Specie/Race', 'attribute': 'specie', 'width': 20},
        '8': {'title': 'Caste', 'attribute': 'caste', 'width': 10},
        '9': {'title': 'Age', 'attribute': 'age', 'width': 10},
        '10': {'title': 'Entrance', 'attribute': 'entrance', 'width': 40},
    }
    ws.cell(column=1, row=1, value='Dramatis Personae')
    character_items = Character.objects.all().order_by('full_name').filter(epic=1, is_visible=True)
    export_header(ws, h)
    cnt = 3
    for c in character_items:
        export_row(ws, h, c, cnt)
        cnt += 1

    # And save everything
    wb.save(filename=dest_filename)


def update_from_xls(filename='dramatis_personae.xlsx'):
    """
      This is not a real 'import', as we only update some refs from the database.
      No isoprod behavior db <-> xls has to be expected here. THIS IS NO RESTORE !!!
    """
    wb = load_workbook(filename)
    ws = wb['Benefices_Afflicitions_References']
    if ws != None:
        cnt = 3
        # bar = BeneficeAfflictionRef.objects.all().delete()
        while True:
            if ws[colrow(1, cnt)].value is None:
                break
            else:
                print(ws[colrow(1, cnt)].value)
                print(ws[colrow(2, cnt)].value)
                print(ws[colrow(3, cnt)].value)
                print(ws[colrow(4, cnt)].value)
                print(ws[colrow(5, cnt)].value)
                print(ws[colrow(6, cnt)].value)
                sheet_id = int(ws[colrow(5, cnt)].value)
                if sheet_id == 0:
                    bar = None
                    print("not found! %d" % (sheet_id))
                else:
                    bar = BeneficeAfflictionRef.objects.get(id=sheet_id)
                    print("found! %d" % (sheet_id))

                if bar is None:
                    bar = BeneficeAfflictionRef(reference=ws[colrow(1, cnt)].value)
                bar.reference = ws[colrow(1, cnt)].value
                bar.source = ws[colrow(6, cnt)].value
                bar.value = int(ws[colrow(2, cnt)].value)
                bar.category = ws[colrow(3, cnt)].value
                if ws[colrow(4, cnt)].value is None:
                    desc = ""
                else:
                    desc = ws[colrow(4, cnt)].value
                bar.description = desc
                bar.save()
                print(bar)
                print("")
                cnt += 1
